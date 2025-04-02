
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

excel_path = 'registro_caracteres.xlsx'

def guardar_codigo(codigo):
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["Fecha/Hora", "Código"])
        wb.save(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), codigo])
    wb.save(excel_path)
